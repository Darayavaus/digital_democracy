"""CLI utility to validate and import DraftLaw rows from an Excel file."""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Dict, Iterator, Tuple

from pydantic import ValidationError
from sqlmodel import Session

from db import engine, init_metadata
from db_model.draft_law import DraftLaw

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "The openpyxl package is required to read Excel files. "
        "Install it with `pip install openpyxl`."
    ) from exc


RowData = Tuple[int, Dict[str, Any]]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import draft laws from an Excel file.")
    parser.add_argument(
        "excel_path",
        type=Path,
        help="Path to the Excel file to import",
    )
    parser.add_argument(
        "--sheet",
        dest="sheet_name",
        help="Worksheet name (defaults to the active sheet)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate and report rows without inserting anything into the database.",
    )
    return parser.parse_args()


def iter_excel_rows(excel_path: Path, sheet_name: str | None) -> Iterator[RowData]:
    workbook = load_workbook(excel_path, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Worksheet '{sheet_name}' not found in file {excel_path}.")
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active

    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return

    headers = [
        (value.strip() if isinstance(value, str) else None)
        for value in header_row
    ]

    for row_number, row_values in enumerate(rows, start=2):
        if all(cell in (None, "") for cell in row_values):
            continue
        row_dict: Dict[str, Any] = {
            column_name: cell_value
            for column_name, cell_value in zip(headers, row_values)
            if column_name
        }
        yield row_number, row_dict


def format_validation_error(exc: ValidationError) -> str:
    parts = []
    for error in exc.errors():
        location = ".".join(str(p) for p in error.get("loc", ()))
        message = error.get("msg", "Invalid value")
        parts.append(f"{location}: {message}" if location else message)
    return "; ".join(parts)


def import_draft_laws(excel_path: Path, sheet_name: str | None, dry_run: bool) -> None:
    excel_path = excel_path.expanduser().resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file {excel_path} was not found.")

    processed = 0
    valid_entries: list[DraftLaw] = []
    warnings: list[Tuple[int, str]] = []

    for row_number, row_data in iter_excel_rows(excel_path, sheet_name):
        processed += 1
        try:
            draft_law = DraftLaw.from_row(row_data)
        except ValidationError as exc:
            warnings.append((row_number, format_validation_error(exc)))
            continue
        except ValueError as exc:
            warnings.append((row_number, str(exc)))
            continue
        valid_entries.append(draft_law)

    print(f"Processed {processed} row(s) from {excel_path}.")
    if warnings:
        print("Warnings for skipped rows:")
        for row_number, message in warnings:
            print(f"  - Row {row_number}: {message}")

    print(f"{len(valid_entries)} row(s) are valid.")

    if not valid_entries:
        print("No valid draft laws to insert.")
        return

    if dry_run:
        print("Dry run enabled; nothing was written to the database.")
        return

    init_metadata()

    with Session(engine) as session:
        session.add_all(valid_entries)
        session.commit()

    print(f"Inserted {len(valid_entries)} draft law(s) into the database.")


def main() -> None:
    args = parse_args()
    import_draft_laws(args.excel_path, args.sheet_name, args.dry_run)


if __name__ == "__main__":
    main()
